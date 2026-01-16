"""
Tulpar Express - Excel Parser Service
Parse Excel files for China warehouse and Bishkek arrival data
"""
from __future__ import annotations

import io
import logging
from dataclasses import dataclass
from typing import List, Optional, BinaryIO

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


@dataclass
class ExcelParcelRow:
    """Parsed parcel row from Excel"""
    client_code: str
    tracking: Optional[str] = None
    weight_kg: Optional[float] = None

    @property
    def is_valid(self) -> bool:
        """Check if row has required data (supports TE-, М-, S-, KG-, N- codes)"""
        if not self.client_code:
            return False
        code = self.client_code.upper().strip()
        # Support multiple code formats: TE-XXXX, М-XXXX, S-XXX, KG-XXX, N-X
        valid_prefixes = ("TE-", "М-", "M-", "S-", "KG-", "N-")
        return any(code.startswith(p) for p in valid_prefixes)


@dataclass
class ExcelParseResult:
    """Result of Excel parsing"""
    rows: List[ExcelParcelRow]
    file_type: str  # "china" or "bishkek"
    errors: List[str]

    @property
    def valid_rows(self) -> List[ExcelParcelRow]:
        return [r for r in self.rows if r.is_valid]

    @property
    def invalid_count(self) -> int:
        return len(self.rows) - len(self.valid_rows)


def detect_file_type(filename: str, sheet_data: List[dict]) -> str:
    """
    Detect Excel file type: china warehouse or bishkek arrival (AC 2.2.4)

    Args:
        filename: Original filename
        sheet_data: First few rows of data

    Returns:
        "china" or "bishkek"
    """
    filename_lower = filename.lower()

    # Check filename for hints
    if "китай" in filename_lower or "china" in filename_lower or "склад" in filename_lower:
        return "china"
    if "бишкек" in filename_lower or "bishkek" in filename_lower or "прибыло" in filename_lower or "оприходов" in filename_lower:
        return "bishkek"

    # Check if data has weight column (bishkek has weight, china doesn't)
    if sheet_data:
        first_row = sheet_data[0]
        headers = [str(k).lower() for k in first_row.keys()]
        if any("вес" in h or "weight" in h or "kg" in h for h in headers):
            return "bishkek"

    # Default to china
    return "china"


def parse_excel(
    file_data: BinaryIO,
    filename: str = "unknown.xlsx",
) -> ExcelParseResult:
    """
    Parse Excel file and extract parcel data (AC 2.3.1-2.3.4)

    Args:
        file_data: File-like object (BytesIO)
        filename: Original filename for type detection

    Returns:
        ExcelParseResult with rows and metadata
    """
    errors = []
    rows = []
    filename_lower = filename.lower()

    # Pre-detect file type from filename
    is_bishkek_file = "оприходов" in filename_lower or "бишкек" in filename_lower or "bishkek" in filename_lower

    try:
        # Load workbook
        wb = load_workbook(file_data, read_only=True, data_only=True)
        ws = wb.active

        if ws is None:
            return ExcelParseResult(rows=[], file_type="unknown", errors=["No active sheet found"])

        # Get all rows
        all_rows = list(ws.iter_rows(values_only=True))

        if len(all_rows) < 1:
            return ExcelParseResult(rows=[], file_type="unknown", errors=["File is empty"])

        # Find actual header row (may not be row 0)
        header_row_idx = find_header_row(all_rows)
        has_headers = header_row_idx >= 0 and is_header_row(all_rows[header_row_idx] if header_row_idx < len(all_rows) else None)

        # Determine column indices based on file structure
        if has_headers:
            logger.info(f"Found headers at row {header_row_idx + 1}")
            headers = [str(h).lower().strip() if h else "" for h in all_rows[header_row_idx]]
            code_col = find_column(headers, ["код", "code", "клиент", "client", "货号"])
            tracking_col = find_column(headers, ["трекинг", "tracking", "номер", "track", "штрих", "快递"])
            weight_col = find_column(headers, ["вес", "weight", "kg", "кг", "重量"])
            data_start_row = header_row_idx + 1
        else:
            # No headers - use positional columns (for Оприходование files)
            logger.info(f"No headers found, using positional columns for {filename}")
            code_col = 0      # Column A = client code
            tracking_col = 1  # Column B = tracking
            weight_col = 2    # Column C = weight
            data_start_row = 0

        # Parse data rows
        for row_num, row in enumerate(all_rows[data_start_row:], start=data_start_row + 1):
            if not row or all(cell is None for cell in row):
                continue

            try:
                # Extract code
                code_value = row[code_col] if code_col < len(row) else None
                if not code_value:
                    continue

                code = normalize_code(code_value)
                if not code:
                    continue

                # Extract tracking
                tracking = None
                if tracking_col is not None and tracking_col < len(row) and row[tracking_col]:
                    tracking = str(row[tracking_col]).strip()
                    # Skip if tracking looks like a code (for header-less files)
                    if tracking and not tracking[0].isdigit() and not tracking.startswith("JT") and not tracking.startswith("YT"):
                        tracking = None

                # Extract weight (handles Russian comma separator)
                weight = None
                if weight_col is not None and weight_col < len(row):
                    weight = parse_weight(row[weight_col])

                rows.append(ExcelParcelRow(
                    client_code=code,
                    tracking=tracking,
                    weight_kg=weight,
                ))

                # For Оприходование files, also check columns I-K for second batch
                if is_bishkek_file and not has_headers and len(row) > 10:
                    code2 = normalize_code(row[8]) if len(row) > 8 and row[8] else None
                    tracking2 = str(row[9]).strip() if len(row) > 9 and row[9] else None
                    weight2 = parse_weight(row[10]) if len(row) > 10 else None

                    if code2:
                        rows.append(ExcelParcelRow(
                            client_code=code2,
                            tracking=tracking2,
                            weight_kg=weight2,
                        ))

            except Exception as e:
                errors.append(f"Ошибка в строке {row_num}: {e}")

        wb.close()

    except Exception as e:
        logger.error(f"Failed to parse Excel: {e}")
        errors.append(f"Ошибка парсинга: {e}")

    # Detect file type
    file_type = detect_file_type(filename, [])
    if rows and any(r.weight_kg for r in rows):
        file_type = "bishkek"

    logger.info(f"Parsed {len(rows)} rows from {filename}, type: {file_type}")

    return ExcelParseResult(
        rows=rows,
        file_type=file_type,
        errors=errors,
    )


def is_header_row(row) -> bool:
    """Check if row looks like a header row (contains keywords)"""
    if not row:
        return False
    header_keywords = ["код", "code", "клиент", "трекинг", "tracking", "вес", "weight", "кг", "kg", "штрих", "货号", "快递", "重量"]
    row_text = " ".join(str(cell).lower() for cell in row if cell)
    matches = sum(1 for kw in header_keywords if kw in row_text)
    return matches >= 2


def find_column(headers: List[str], keywords: List[str]) -> Optional[int]:
    """Find column index by keywords"""
    for idx, header in enumerate(headers):
        for keyword in keywords:
            if keyword in header:
                return idx
    return None


def find_header_row(all_rows: List[tuple], max_scan: int = 10) -> int:
    """
    Find the actual header row by scanning first N rows for keywords.
    Real Excel files often have title rows before actual headers.

    Returns:
        Row index (0-based) where headers are found
    """
    header_keywords = ["код", "code", "клиент", "трекинг", "tracking", "вес", "weight", "кг", "kg", "штрих"]

    for row_idx, row in enumerate(all_rows[:max_scan]):
        if row is None:
            continue
        # Convert row to lowercase strings for matching
        row_text = " ".join(str(cell).lower() for cell in row if cell)

        # Count how many keywords match
        matches = sum(1 for kw in header_keywords if kw in row_text)

        # If we find 2+ keywords, this is likely the header row
        if matches >= 2:
            return row_idx

    # Default to row 0 if no headers found
    return 0


def parse_weight(value) -> Optional[float]:
    """
    Parse weight value, handling Russian comma decimal separator.

    Examples:
        0.5 -> 0.5
        "0,3" -> 0.3
        "1.5" -> 1.5
    """
    if value is None:
        return None

    if isinstance(value, (int, float)):
        return float(value)

    try:
        # Convert string, replacing comma with dot
        str_value = str(value).strip().replace(",", ".")
        return float(str_value)
    except (ValueError, TypeError):
        return None


def normalize_code(value) -> str:
    """
    Normalize client code preserving original format.

    Supports: TE-XXXX, S-XXX, KG-XXX, N-X, М-XXXX

    Examples:
        "TE-5001" -> "TE-5001"
        "te-5001" -> "TE-5001"
        "S-601" -> "S-601"
        "KG-127" -> "KG-127"
        "N-7" -> "N-7"
        "5001" -> "TE-5001" (default to TE- if no prefix)
    """
    if value is None:
        return ""
    value = str(value).strip().upper()

    # Normalize Russian ТЕ to Latin TE
    value = value.replace("ТЕ-", "TE-").replace("ТЕ", "TE-")

    # Known prefixes - preserve them
    known_prefixes = ("TE-", "S-", "KG-", "N-", "М-", "M-")
    for prefix in known_prefixes:
        if value.startswith(prefix):
            return value

    # If no known prefix but has digits, default to TE-
    digits = "".join(filter(str.isdigit, value))
    if digits:
        return f"TE-{int(digits):04d}"

    return value
